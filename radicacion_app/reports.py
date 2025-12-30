from django.views.generic import TemplateView
from django.http.response import HttpResponse
from .models import RadicacionRecibidos, RadicacionInternos, RadicacionEnviados, RadicadosRecibidosPqrsd, RadicadosEnviadosPqrsd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment



class ReporterExcelRadicadosRecibidos(TemplateView):
    def __init__(self, queryset = None):
        # Utiliza el queryset pasado en el constructor o todos los registros si no se pasa
        self.queryset = queryset or RadicacionRecibidos.objects.all()
    
    def get(self, request, *args, **kwargs):
        # Aqui utilizamos el queryset filtrado en lugar de obtener todos los registros
        radicados = self.queryset        
        
        # Crea el archivo excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Radicados Recibidos"
        
        # Estilos
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="C4D01C", end_color="C4D01C", fill_type="solid")
        center_alignment = Alignment(horizontal="center")
        
        ws["B1"] = "REPORTE DE RADICADOS RECIBIDOS"
        ws.merge_cells("B1:L1")
        ws["B1"].font = title_font
        ws["B1"].alignment = center_alignment
        
        headers = ["FECHA RADICADO", "No RADICADO", "MEDIO INGRESO", "TIPO COMUNICACION", "TIPO DOCUMENTO", "No RADICADO DE LLEGADA",
                "ENTIDAD", "ASUNTO", "ANEXOS", "RESPONSABLE", "OFICINA", "REQUIERE RESPUESTA", "TIEMPO RESPUESTA", "FECHA MAXIMA DE RESPUESTA",
                "OBSERVACIONES", "RADICO", "RESPUESTA RADICADO INTERNO", "ESTADO RESPUESTA", "FECHA RESPUESTA", "MEDIO DE RESPUESTA"]
        
        for col_num, header in enumerate(headers, 2):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        cont = 4

        for radicado in radicados:
            

            # Agregar datos al Excel
            ws.cell(row=cont, column=2).value = str(radicado.fecha_radicacion)
            ws.cell(row=cont, column=3).value = str(radicado.id)
            ws.cell(row=cont, column=4).value = str(radicado.medio_ingreso)
            ws.cell(row=cont, column=5).value = str(radicado.tipo_comunicacion)
            ws.cell(row=cont, column=6).value = str(radicado.tipo_documento)
            ws.cell(row=cont, column=7).value = str(radicado.num_rad_llegada)
            ws.cell(row=cont, column=8).value = str(radicado.entidad)
            ws.cell(row=cont, column=9).value = str(radicado.asunto)
            ws.cell(row=cont, column=10).value = str(radicado.anexos)
            ws.cell(row=cont, column=11).value = str(radicado.responsable_por_responder.nombre)
            ws.cell(row=cont, column=12).value = str(radicado.oficina)
            ws.cell(row=cont, column=13).value = str(radicado.requiere_respuesta)
            ws.cell(row=cont, column=14).value = str(radicado.tiempo_respuesta)
            ws.cell(row=cont, column=15).value = str(radicado.fecha_maxima_respuesta)
            ws.cell(row=cont, column=16).value = str(radicado.observaciones)
            ws.cell(row=cont, column=17).value = str(radicado.radicador)
            ws.cell(row=cont, column=18).value = str(radicado.respuesta_rad_interno)
            ws.cell(row=cont, column=19).value = str(radicado.estado_respuesta)
            ws.cell(row=cont, column=20).value = str(radicado.fecha_respuesta)
            ws.cell(row=cont, column=21).value = str(radicado.medio_respuesta)

            cont += 1

        
        # Ajustar el ancho de las columnas
        for col_num in range(1, len(headers) + 1):
            max_length = 0
            column_letter = ws.cell(row=3, column=col_num).column_letter
            for cell in ws[column_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
                    

        file_name = "Reporte_radicados_recibidos.xlsx"
        response = HttpResponse(content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        content = "attachment; filename = {0}".format(file_name)
        response["Content-Disposition"] = content
        wb.save(response)
        return response




class ReporterExcelRadicadosEnviados(TemplateView):
    def __init__(self, queryset = None):
        # Utiliza el queryset pasado en el constructor o todos los registros si no se pasa
        self.queryset = queryset or RadicacionEnviados.objects.all()
    
    def get(self, request, *args, **kwargs):
        # Aqui utilizamos el queryset filtrado en lugar de obtener todos los registros
        radicados = self.queryset        
        
        # Crea el archivo excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Radicados Enviados"
        
        # Estilos
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="C4D01C", end_color="C4D01C", fill_type="solid")
        center_alignment = Alignment(horizontal="center")
        
        ws["B1"] = "REPORTE DE RADICADOS ENVIADOS"
        ws.merge_cells("B1:L1")
        ws["B1"].font = title_font
        ws["B1"].alignment = center_alignment
        
        headers = ["FECHA RADICADO", "No RADICADO", "MEDIO ENVIO", "TIPO COMUNICACION", "TIPO DOCUMENTO", "ENTIDAD", "ASUNTO", 
                   "No RADICADO DE LLEGADA", "No RADICADO INTERNO", "ANEXOS", "ENVIADO POR", "OFICINA", "OBSERVACIONES", "RADICO", 
                   "REQUIERE RESPUESTA", "TIEMPO RESPUESTA", "FECHA MAXIMA RESPUESTA", "FECHA RESPUESTA", "RECIBIDO"]
        
        for col_num, header in enumerate(headers, 2):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        cont = 4

        for radicado in radicados:
            

            # Agregar datos al Excel
            ws.cell(row=cont, column=2).value = str(radicado.fecha_radicacion)
            ws.cell(row=cont, column=3).value = str(radicado.id)
            ws.cell(row=cont, column=4).value = str(radicado.medio_envio)
            ws.cell(row=cont, column=5).value = str(radicado.tipo_comunicacion)
            ws.cell(row=cont, column=6).value = str(radicado.tipo_documento)
            ws.cell(row=cont, column=7).value = str(radicado.entidad)
            ws.cell(row=cont, column=8).value = str(radicado.asunto)
            ws.cell(row=cont, column=9).value = str(radicado.num_rad_llegada)
            ws.cell(row=cont, column=10).value = str(radicado.num_rad_interno)
            ws.cell(row=cont, column=11).value = str(radicado.anexos)
            ws.cell(row=cont, column=12).value = str(radicado.enviado_por.nombre)
            ws.cell(row=cont, column=13).value = str(radicado.oficina)
            ws.cell(row=cont, column=14).value = str(radicado.observaciones)
            ws.cell(row=cont, column=15).value = str(radicado.radicador)
            ws.cell(row=cont, column=16).value = str(radicado.requiere_respuesta)
            ws.cell(row=cont, column=17).value = str(radicado.tiempo_respuesta)
            ws.cell(row=cont, column=18).value = str(radicado.fecha_maxima_respuesta)
            ws.cell(row=cont, column=19).value = str(radicado.fecha_respuesta)
            ws.cell(row=cont, column=20).value = str(radicado.recibido)

            cont += 1

        
        # Ajustar el ancho de las columnas
        for col_num in range(1, len(headers) + 1):
            max_length = 0
            column_letter = ws.cell(row=3, column=col_num).column_letter
            for cell in ws[column_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
                    

        file_name = "Reporte_radicados_enviados.xlsx"
        response = HttpResponse(content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        content = "attachment; filename = {0}".format(file_name)
        response["Content-Disposition"] = content
        wb.save(response)
        return response





class ReporterExcelRadicadosInternos(TemplateView):
    def __init__(self, queryset = None):
        # Utiliza el queryset pasado en el constructor o todos los registros si no se pasa
        self.queryset = queryset or RadicacionInternos.objects.all()
    
    def get(self, request, *args, **kwargs):
        # Aqui utilizamos el queryset filtrado en lugar de obtener todos los registros
        radicados = self.queryset        
        
        # Crea el archivo excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Radicados Internos"
        
        # Estilos
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="C4D01C", end_color="C4D01C", fill_type="solid")
        center_alignment = Alignment(horizontal="center")
        
        ws["B1"] = "REPORTE DE RADICADOS INTERNOS"
        ws.merge_cells("B1:L1")
        ws["B1"].font = title_font
        ws["B1"].alignment = center_alignment
        
        headers = ["FECHA RADICADO", "No RADICADO", "MEDIO INGRESO", "TIPO COMUNICACION", "TIPO DOCUMENTO", "No RADICADO DE LLEGADA",
                "ENTIDAD", "ASUNTO", "ANEXOS", "RESPONSABLE", "OFICINA", "REQUIERE RESPUESTA", "TIEMPO RESPUESTA", "FECHA MAXIMA DE RESPUESTA",
                "OBSERVACIONES", "RADICO", "RESPUESTA RADICADO INTERNO", "ESTADO RESPUESTA", "FECHA RESPUESTA", "MEDIO DE RESPUESTA"]
        
        for col_num, header in enumerate(headers, 2):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        cont = 4

        for radicado in radicados:
            

            # Agregar datos al Excel
            ws.cell(row=cont, column=2).value = str(radicado.fecha_radicacion)
            ws.cell(row=cont, column=3).value = str(radicado.id)
            ws.cell(row=cont, column=4).value = str(radicado.medio_ingreso)
            ws.cell(row=cont, column=5).value = str(radicado.tipo_comunicacion)
            ws.cell(row=cont, column=6).value = str(radicado.tipo_documento)
            ws.cell(row=cont, column=7).value = str(radicado.num_rad_llegada)
            ws.cell(row=cont, column=8).value = str(radicado.entidad)
            ws.cell(row=cont, column=9).value = str(radicado.asunto)
            ws.cell(row=cont, column=10).value = str(radicado.anexos)
            ws.cell(row=cont, column=11).value = str(radicado.responsable_por_responder.nombre)
            ws.cell(row=cont, column=12).value = str(radicado.oficina)
            ws.cell(row=cont, column=13).value = str(radicado.requiere_respuesta)
            ws.cell(row=cont, column=14).value = str(radicado.tiempo_respuesta)
            ws.cell(row=cont, column=15).value = str(radicado.fecha_maxima_respuesta)
            ws.cell(row=cont, column=16).value = str(radicado.observaciones)
            ws.cell(row=cont, column=17).value = str(radicado.radicador)
            ws.cell(row=cont, column=18).value = str(radicado.respuesta_rad_interno)
            ws.cell(row=cont, column=19).value = str(radicado.estado_respuesta)
            ws.cell(row=cont, column=20).value = str(radicado.fecha_respuesta)
            ws.cell(row=cont, column=21).value = str(radicado.medio_respuesta)

            cont += 1

        
        # Ajustar el ancho de las columnas
        for col_num in range(1, len(headers) + 1):
            max_length = 0
            column_letter = ws.cell(row=3, column=col_num).column_letter
            for cell in ws[column_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
                    

        file_name = "Reporte_radicados_internos.xlsx"
        response = HttpResponse(content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        content = "attachment; filename = {0}".format(file_name)
        response["Content-Disposition"] = content
        wb.save(response)
        return response



class ReporterExcelPqrsdRecibidos(TemplateView):
    def __init__(self, queryset = None):
        # Utiliza el queryset pasado en el constructor o todos los registros si no se pasa
        self.queryset = queryset or RadicadosRecibidosPqrsd.objects.all()
    
    def get(self, request, *args, **kwargs):
        # Aqui utilizamos el queryset filtrado en lugar de obtener todos los registros
        radicados = self.queryset        
        
        # Crea el archivo excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Pqrsd Recibidos"
        
        # Estilos
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="C4D01C", end_color="C4D01C", fill_type="solid")
        center_alignment = Alignment(horizontal="center")
        
        ws["B1"] = "REPORTE DE PQRSD RECIBIDOS"
        ws.merge_cells("B1:L1")
        ws["B1"].font = title_font
        ws["B1"].alignment = center_alignment
        
        headers = ["FECHA RADICADO", "No RADICADO", "RADICO", "TIPO RADICADO", "UNIDAD DE NEGOCIO", "FECHA RECIBIDO USUARIO",
                   "FECHA ASIGNACION TRASLADO", "RADICADO RECIBIDO", "CANAL DE RECEPCION", "TIPO PETICION", "NOMBRE REMITENTE",
                   "TELEFONO", "DIRECCION", "CORREO ELECTRONICO", "ASUNTO", "FECHA EVENTO", "HORA EVENTO", "LUGAR EVENTO", "SERIAL O PLACA",
                   "RUTA", "DESCRIPCION", "ASIGNADO A", "VENCIMIENTO INTERNO", "VENCIMIENTO POR LEY", "FECHA DE CIERRE", "CULPABILIDAD",
                   "OPERADOR","OBSERVACIONES" ]
        
        for col_num, header in enumerate(headers, 2):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        cont = 4

        for radicado in radicados:
            

            # Agregar datos al Excel
            ws.cell(row=cont, column=2).value = str(radicado.fecha_radicacion)
            ws.cell(row=cont, column=3).value = str(radicado.id)
            ws.cell(row=cont, column=4).value = str(radicado.radicador)
            ws.cell(row=cont, column=5).value = str(radicado.tipo_radicado)
            ws.cell(row=cont, column=6).value = str(radicado.unidad_negocio)
            ws.cell(row=cont, column=7).value = str(radicado.fecha_recibido_usuario)
            ws.cell(row=cont, column=8).value = str(radicado.fecha_asignacion_traslado)
            ws.cell(row=cont, column=9).value = str(radicado.radicado_recibido)
            ws.cell(row=cont, column=10).value = str(radicado.canal_recepcion)
            ws.cell(row=cont, column=11).value = str(radicado.tipo_peticion)
            ws.cell(row=cont, column=12).value = str(radicado.nombre_remitente)
            ws.cell(row=cont, column=13).value = str(radicado.telefono)
            ws.cell(row=cont, column=14).value = str(radicado.direccion)
            ws.cell(row=cont, column=15).value = str(radicado.email)
            ws.cell(row=cont, column=16).value = str(radicado.asunto)
            ws.cell(row=cont, column=17).value = str(radicado.fecha_evento)
            ws.cell(row=cont, column=18).value = str(radicado.hora_evento)
            ws.cell(row=cont, column=19).value = str(radicado.lugar_evento)
            ws.cell(row=cont, column=20).value = str(radicado.serial_o_placa)
            ws.cell(row=cont, column=21).value = str(radicado.ruta)
            ws.cell(row=cont, column=22).value = str(radicado.descripcion)
            ws.cell(row=cont, column=23).value = str(radicado.asignado_a)
            ws.cell(row=cont, column=24).value = str(radicado.vencimiento_interno)
            ws.cell(row=cont, column=25).value = str(radicado.vencimiento_por_ley)
            ws.cell(row=cont, column=26).value = str(radicado.fecha_cierre)
            ws.cell(row=cont, column=27).value = str(radicado.culpabilidad)
            ws.cell(row=cont, column=28).value = str(radicado.operador)
            ws.cell(row=cont, column=29).value = str(radicado.observaciones)
            cont += 1

        
        # Ajustar el ancho de las columnas
        for col_num in range(1, len(headers) + 1):
            max_length = 0
            column_letter = ws.cell(row=3, column=col_num).column_letter
            for cell in ws[column_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
                    

        file_name = "Reporte_pqrsd_recibidos.xlsx"
        response = HttpResponse(content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        content = "attachment; filename = {0}".format(file_name)
        response["Content-Disposition"] = content
        wb.save(response)
        return response


class ReporterExcelPqrsdEnviados(TemplateView):
    def __init__(self, queryset = None):
        # Utiliza el queryset pasado en el constructor o todos los registros si no se pasa
        self.queryset = queryset or RadicadosEnviadosPqrsd.objects.all()
    
    def get(self, request, *args, **kwargs):
        # Aqui utilizamos el queryset filtrado en lugar de obtener todos los registros
        radicados = self.queryset        
        
        # Crea el archivo excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Pqrsd Enviados"
        
        # Estilos
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="C4D01C", end_color="C4D01C", fill_type="solid")
        center_alignment = Alignment(horizontal="center")
        
        ws["B1"] = "REPORTE DE PQRSD ENVIADOS"
        ws.merge_cells("B1:L1")
        ws["B1"].font = title_font
        ws["B1"].alignment = center_alignment
        
        headers = ["FECHA RADICADO", "No RADICADO", "RADICO", "TIPO RADICADO", "ASUNTO", "RADICADO ASOCIADO", "DESTINATARIO"]
        
        for col_num, header in enumerate(headers, 2):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        cont = 4

        for radicado in radicados:
            

            # Agregar datos al Excel
            ws.cell(row=cont, column=2).value = str(radicado.fecha_radicacion)
            ws.cell(row=cont, column=3).value = str(radicado.id)
            ws.cell(row=cont, column=4).value = str(radicado.radicador)
            ws.cell(row=cont, column=5).value = str(radicado.tipo_radicado)
            ws.cell(row=cont, column=6).value = str(radicado.asunto)
            ws.cell(row=cont, column=7).value = str(radicado.radicado_asociado)
            ws.cell(row=cont, column=8).value = str(radicado.destinatario)

            cont += 1

        
        # Ajustar el ancho de las columnas
        for col_num in range(1, len(headers) + 1):
            max_length = 0
            column_letter = ws.cell(row=3, column=col_num).column_letter
            for cell in ws[column_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
                    

        file_name = "Reporte_pqrsd_enviados.xlsx"
        response = HttpResponse(content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        content = "attachment; filename = {0}".format(file_name)
        response["Content-Disposition"] = content
        wb.save(response)
        return response