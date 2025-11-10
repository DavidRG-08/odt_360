from django.views.generic import TemplateView
from django.http.response import HttpResponse
from .models import SolicitudRuta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.contrib.auth.models import User

class ReporterExcelRutas(TemplateView):
    def __init__(self, queryset = None):
        # Utiliza el queryset pasado en el constructor o todos los registros si no se pasa
        self.queryset = queryset or SolicitudRuta.objects.all()
        
    
    def get(self, request, *args, **kwargs):
        # Aqui utilizamos el queryset filtrado en lugar de obtener todos los registros
        rutas_reg = self.queryset        
        
        # Crea el archivo excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Rutas Solicitadas"
        
        # Estilos
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="C4D01C", end_color="C4D01C", fill_type="solid")
        center_alignment = Alignment(horizontal="center")
        
        ws["B1"] = "REPORTE RUTAS SOLICITADAS"
        ws.merge_cells("B1:L1")
        ws["B1"].font = title_font
        ws["B1"].alignment = center_alignment

        headers = ["FECHA", "OPERADOR", "NOMBRE", "TELEFONO", "TURNO", "FECHA DE RECOGIDA",
                "LOCALIDAD", "BARRIO", "UBICACION", "RUTA", "DIRECCION", "ESTADO"]
        
        for col_num, header in enumerate(headers, 2):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        cont = 4

        # Modificacion para el event stop code

        for ruta in rutas_reg:
            
            # try:
            #     user = User.objects.get(username=ruta.operador)
            #     nombre = f"{user.first_name} {user.last_name}"
            # except User.DoesNotExist:
            #     nombre = "Desconocido"

            # Agregar datos al Excel
            ws.cell(row=cont, column=2).value = str(ruta.fecha_solicitud)
            ws.cell(row=cont, column=3).value = str(ruta.operador)
            # ws.cell(row=cont, column=4).value = nombre
            ws.cell(row=cont, column=5).value = str(ruta.telefono)
            ws.cell(row=cont, column=6).value = str(ruta.turno)
            ws.cell(row=cont, column=7).value = str(ruta.fecha_recogida)
            ws.cell(row=cont, column=8).value = str(ruta.localidad)
            ws.cell(row=cont, column=9).value = str(ruta.barrio)
            ws.cell(row=cont, column=10).value = str(ruta.ubicacion)
            ws.cell(row=cont, column=11).value = str(ruta.ruta)
            ws.cell(row=cont, column=12).value = str(ruta.direccion)
            ws.cell(row=cont, column=13).value = str("Activa" if ruta.estado else "Cancelada")

            cont += 1

        
        # Ajustar el ancho de las columnas
        for col_num in range(1, len(headers) + 1):
            max_length = 0
            column_letter = ws.cell(row=3, column=col_num).column_letter
            for cell in ws[column_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
                    

        file_name = "Reporte_rutas_solicitadas.xlsx"
        response = HttpResponse(content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        content = "attachment; filename = {0}".format(file_name)
        response["Content-Disposition"] = content
        wb.save(response)
        return response
    

