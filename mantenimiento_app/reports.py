from django.views.generic import TemplateView
from django.http.response import HttpResponse
from .models import OrdenAlistamiento
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

class ReporterExcelOrdenesAlistamiento(TemplateView):
    def __init__(self, queryset = None):
        # Utiliza el queryset pasado en el constructor o todos los registros si no se pasa
        self.queryset = queryset or OrdenAlistamiento.objects.all()
    
    def get(self, request, *args, **kwargs):
        # Aqui utilizamos el queryset filtrado en lugar de obtener todos los registros
        ordenes = self.queryset        
        
        # Crea el archivo excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Ordenes Alistamiento"
        
        # Estilos
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="C4D01C", end_color="C4D01C", fill_type="solid")
        center_alignment = Alignment(horizontal="center")
        
        ws["B1"] = "REPORTE ORDENES ALISTAMIENTO"
        ws.merge_cells("B1:L1")
        ws["B1"].font = title_font
        ws["B1"].alignment = center_alignment
        
        headers = ["FECHA CREACION", "No ORDEN", "ESTADO", "TECNICO MTTO", "VEHICULO", "FECHA INICIO ORDEN",
                "FECHA FIN ORDEN", "TIEMPO ALISTAMIENTO", "NOVEDADES"]
        
        for col_num, header in enumerate(headers, 2):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        cont = 4

        for orden in ordenes:
            

            # Agregar datos al Excel
            ws.cell(row=cont, column=2).value = str(orden.fecha_creacion_orden)
            ws.cell(row=cont, column=3).value = str(orden.id)
            ws.cell(row=cont, column=4).value = str(orden.estado)
            ws.cell(row=cont, column=5).value = str(orden.user)
            ws.cell(row=cont, column=6).value = str(orden.vehiculo)
            ws.cell(row=cont, column=7).value = str(orden.fecha_inicio)
            ws.cell(row=cont, column=8).value = str(orden.fecha_fin)
            ws.cell(row=cont, column=9).value = str(orden.tiempo_alistamiento)
            ws.cell(row=cont, column=10).value = str(orden.novedades)
            # ws.cell(row=cont, column=12).value = str("Activa" if orden.estado else "Cancelada")

            cont += 1

        
        # Ajustar el ancho de las columnas
        for col_num in range(1, len(headers) + 1):
            max_length = 0
            column_letter = ws.cell(row=3, column=col_num).column_letter
            for cell in ws[column_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
                    

        file_name = "Reporte_ordenes_de_alistamiento.xlsx"
        response = HttpResponse(content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        content = "attachment; filename = {0}".format(file_name)
        response["Content-Disposition"] = content
        wb.save(response)
        return response
    

